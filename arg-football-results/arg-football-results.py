# Usa python3
import requests,datetime,openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, Style, PatternFill, Border, Side, Alignment, Protection
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart import BarChart, Series, Reference

from bs4 import BeautifulSoup

from match import Match

################################
# Constants to adapt each time #
################################

years = [2018,2019]
teams_year = {2018:26,2019:24}
start_year = {2018:1342,2019:2006}
finish_year = {2018:-195,2019:-135}
rounds_year = {2018:25,2019:13}
wikipedia_fields = 6

year_month = {year:{'january':year+1,'enero':year+1, 'february':year+1,'febrero':year+1, 
						  'march':year+1,'marzo':year+1, 'april':year+1,'abril':year+1,	'may':year+1,'mayo':year+1,
						  'june':year+1,'junio':year+1, 'july':year,'julio':year, 'august':year,'agosto':year,
						  'september':year,'septiembre':year,'setiembre':year, 'october':year,'octubre':year, 
						   'november':year,'noviembre':year, 'december':year,'diciembre':year} for year in years}

index_month = {year:{'january':1,'enero':1, 'february':2,'febrero':2, 'march':3,'marzo':3, 'april':4,'abril':4,
					'may':5,'mayo':5,'june':6,'junio':6, 'july':7,'julio':7, 'august':8,'agosto':8,
					'september':9,'septiembre':9,'setiembre':9, 'october':10,'octubre':10, 
					'november':11,'noviembre':11, 'december':12,'diciembre':12} for year in years}
					
index_datetime_day = {0:'monday', 1:'tuesday', 2:'wednesday', 3:'thursday', 4:'friday', 5:'saturday', 6:'sunday'}
index_datetime_spanish_day = {0:'lunes', 1:'martes', 2:'miercoles', 3:'jueves', 4:'viernes', 5:'sabado', 6:'domingo'}

colors = {	'black' :'FF000000','white' : 'FFFFFFFF', 
			'light_red' : 'FFFF4D4D', 'dark_red' : 'FFB30000', 'red' : 'FFFF3333',
			'light_blue' : 'FF4D94FF', 'dark_blue' : 'FF0066CC', 'blue' : 'FF3385FF', 
			'light_green' : 'FF66FF33', 'dark_green' : 'FF009900', 'green' : 'FF47D147',
			'light_orange' : 'FFFFA366', 'dark_orange' : 'FFCC5200', 'orange' : 'FFFF8533',
			'light_yellow' : 'FFFFFF66', 'dark_yellow' : 'FF666600', 'yellow' : 'FFB3B300',
			'light_violet' : 'FFB380FF', 'dark_violet' : 'FF884DFF', 'violet' : 'FFAA80FF'}

def get_data_from_year(year):
	wiki = "https://es.wikipedia.org/wiki/Campeonato_de_Primera_División_" + str(year) +"-" + str(year+1)[2:] + "_(Argentina)"
	file_name = 'horarios_superliga_' + str(year) + '-' + str(year+1)[2:] + '.xlsx'
	website_url = requests.get(wiki).text
	soup = BeautifulSoup(website_url,'lxml')
	teams = teams_year[year]
	rounds = rounds_year[year]
	rows = (teams//2)*rounds
	return (soup,teams,rounds,rows,file_name)

def get_matches_from_wikipedia(soup,teams,rounds,rows,year):

    matches_raw = [['' for x in range(wikipedia_fields)] + [row//(teams//2)+1] for row in range(rows)]
    i = 0
    start = start_year[year]   # Search on the table watching HTML from wikipedia, aided by bs4
    finish = finish_year[year] # Search on the table watching HTML from wikipedia, aided by bs4
    
		
    for x in soup.select('table tbody tr td')[start:finish]:
        while(i < wikipedia_fields*rows and matches_raw[i//wikipedia_fields][i%wikipedia_fields] != ''):
            i += 1
        if i < wikipedia_fields*rows:
            field = x.getText().strip().split('[')[0]
            if field == '':
                field = "Entry without value"
            if i % wikipedia_fields == 4:  # Date
                day_month = list(map(lambda s : s.strip().lower(),field.split(' de ')))
                field = datetime.date(year_month[year][day_month[1]], index_month[year][day_month[1]], int(day_month[0]))
            k = 1
            if x.get('rowspan') != None:
                k = int(x.get('rowspan'))
            for j in range(k):
                matches_raw[i//wikipedia_fields+j][i%wikipedia_fields] = field
    
    match_list = []
    for match in matches_raw:
        hours, minutes = list(map(int, match[5].split(':')))
        home_goals,away_goals = map(lambda s : s.strip(), match[1].split('-'))
        home,away = match[0],match[2]
        stadium = match[3]
        date = datetime.datetime(match[4].year, match[4].month, match[4].day, hours, minutes)
        round_slot = match[6]
        match_list.append(Match(date,round_slot,stadium,home,away,home_goals,away_goals,False))
    
    return match_list

def print_value(sheet, r, c, value, is_date, is_time, top_border, background_color, text_color, is_bold):
	sheet.cell(row = r, column = c).value = value
	if is_date:
		sheet.cell(row = r, column = c).number_format = 'dd/mm/yyyy'
	if is_time:
		sheet.cell(row = r, column = c).number_format = 'hh:mm'
	
	sheet.cell(row = r, column = c).fill = PatternFill(start_color=colors[background_color],end_color=colors[background_color],fill_type='solid')
	sheet.cell(row = r, column = c).font = Font(name='Calibri',size=11,bold=is_bold, color = colors[text_color])	
	sheet.cell(row = r, column = c).alignment = Alignment(horizontal='center',vertical='center')
	if top_border:
		sheet.cell(row = r, column = c).border = Border(left=Side(border_style='thin', color=colors['black']),right=Side(border_style='thin',color=colors['black']),top=Side(border_style='thick',color=colors['black']),bottom=Side(border_style='thin',color=colors['black']))
	else:
		sheet.cell(row = r, column = c).border = Border(left=Side(border_style='thin', color=colors['black']),right=Side(border_style='thin',color=colors['black']),top=Side(border_style='thin',color=colors['black']),bottom=Side(border_style='thin',color=colors['black']))
		

def format_sheet(sheet):
	sheet.freeze_panes = 'A2'
	for c in 'EFGHIJKLMNOPQRSTUVWXYZ':
		sheet.column_dimensions[c].width = 25
	for c in 'BCD':
		sheet.column_dimensions[c].width = 15
	for c in 'A':
		sheet.column_dimensions[c].width = 10
	
    
def generate_data_sheet(file_name,match_list):
	wb = openpyxl.Workbook()

	sheet = wb.get_active_sheet()
	sheet.title = 'Horarios Superliga'

	sheet_columns = ["Fecha","Calendario","Dia Semana","Horario","Local","Visitante"]
	color_columns = [["light_yellow","light_blue","light_red","light_green","light_violet","light_orange"],
					 ["yellow","blue","red","green","violet","orange"],
					 ["dark_yellow","dark_blue","dark_red","dark_green","dark_violet","dark_orange"]]
	format_sheet(sheet)
	
	r = 1
	for c in range(len(sheet_columns)):
		print_value(sheet,r,c+1,sheet_columns[c],False,False,True,color_columns[2][c],'white',True)

	last_round = 0
	last_day = -1
	ROWS_BETWEEN_ROUNDS = 10
	for match in match_list:
		
		if match.round_slot != last_round:
			last_round = match.round_slot
			last_day = match.date.weekday()
			r += ROWS_BETWEEN_ROUNDS

		row_values = [match.round_slot,match.date, match.date, index_datetime_spanish_day[match.date.weekday()].upper(),match.home.upper(),match.away.upper()]
		top_border = match.date.weekday() != last_day
		last_day = match.date.weekday()
		
		for c in range(len(row_values)):
			background_color = color_columns[match.round_slot%2][c]
			print_value(sheet,r,c+1,row_values[c], c == 1, c == 2, top_border, background_color, 'black',False)
		r += 1
		
	wb.save(file_name)
       
def main():
	for year in years:
		soup,teams,rounds,rows,file_name = get_data_from_year(year)
		match_list = sorted(get_matches_from_wikipedia(soup,teams,rounds,rows,year))
		generate_data_sheet(file_name,match_list)
		

if __name__ == '__main__':
	main()
	
	
    
        
    # TO DO: Hacer mejor con pandas.
	#~ print("\n\n\n#######################\n# Superliga " + str(year) + "-" + str(year+1) + " #\n#######################\n")
    
    # nombres_equipos = equipos = sorted(list(set([p[0] for p in partidos])))
    # for eq in nombres_equipos:
    #     dia_horario_local = {}
    #     dia_horario_visitante = {}
    #     dia_horario_total = {}
    #     for p in partidos:
    #         tupla_horario = (p[4].weekday(), p[5])
    #         if p[0] == eq or p[2] == eq:
    #             if not tupla_horario in dia_horario_total.keys():
    #                 dia_horario_total[tupla_horario] = 0
    #                 dia_horario_local[tupla_horario] = 0
    #                 dia_horario_visitante[tupla_horario] = 0
    #             dia_horario_total[tupla_horario] += 1
    #             if p[0] == eq:
    #                 dia_horario_local[tupla_horario] += 1
    #             else:
    #                 dia_horario_visitante[tupla_horario] += 1
    #
    #     suma_partidos = 0
    #     # print("\n#############\n " + eq + "\n#############\n")
    #     # print("TOTAL,LOCAL,VISITANTE:\n")
    #     for c,h in sorted([(dia_horario_total[x],x) for x in dia_horario_total.keys()], reverse = True):
    #         suma_partidos += c
    #         # print(indice_datetime_a_dia[h[0]] + " " + h[1], "--> ", str(c) + ", " + str(dia_horario_local[h]) + ", " + str(dia_horario_visitante[h]))
    #     assert(suma_partidos == fechas)



